from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / 'Data UMKM Online Shop Kabupaten Mojokerto.xlsx'
OUTPUT_PATH = ROOT / 'ui' / 'ui_seed_data.json'


def safe_value(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def as_float(value):
    if value in (None, ''):
        return None
    text = str(value).strip().replace('.', '').replace(',', '.')
    try:
        return float(text)
    except ValueError:
        return None


def parse_rw_rt(nmsls: str | None) -> tuple[str | None, str | None]:
    """Parse RT and RW from nmsls field like 'RT 001 RW 003 DUSUN WONOSARI'"""
    if not nmsls:
        return None, None
    
    rw_match = re.search(r'RW\s+(\d+)', str(nmsls), re.IGNORECASE)
    rt_match = re.search(r'RT\s+(\d+)', str(nmsls), re.IGNORECASE)
    
    rw = rw_match.group(1) if rw_match else None
    rt = rt_match.group(1) if rt_match else None
    
    return rt, rw


def normalize_coords(lat, lon):
    """Attempt to normalize latitude/longitude values to valid decimal degrees.
    Returns (lat, lon) as floats or (None, None) if cannot normalize.
    Tries common scaling factors and swapping if needed.
    """
    try:
        if lat is None or lon is None:
            return None, None
        lat = float(lat)
        lon = float(lon)
    except Exception:
        return None, None

    def is_valid(a, b):
        return abs(a) <= 90 and abs(b) <= 180

    # If already valid, return
    if is_valid(lat, lon):
        return lat, lon

    # Try dividing by common factors
    factors = [1e6, 1e7, 1e5, 1e4, 1e3, 100, 10]
    for f in factors:
        la = lat / f
        lo = lon / f
        if is_valid(la, lo):
            return la, lo

    # Try swapping lat/lon and repeat
    for f in factors:
        la = lon / f
        lo = lat / f
        if is_valid(la, lo):
            return la, lo

    # Last resort: try integer shifts (remove extra digits by dividing by 10 repeatedly)
    la, lo = lat, lon
    for _ in range(7):
        la /= 10
        lo /= 10
        if is_valid(la, lo):
            return la, lo

    return None, None


def build_seed() -> dict:
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)

    gmaps_ws = wb['Master_GoogleMaps']
    tokped_ws = wb['Master_Tokopedia']

    gmaps_header = list(next(gmaps_ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    tokped_header = list(next(tokped_ws.iter_rows(min_row=1, max_row=1, values_only=True)))

    gmaps_idx = {name: i for i, name in enumerate(gmaps_header)}
    tokped_idx = {name: i for i, name in enumerate(tokped_header)}

    gmaps_rows = []
    tokped_rows = []
    kecamatan_set = set()
    desa_set = set()
    desa_by_kecamatan = defaultdict(set)
    rw_set_by_desa = defaultdict(set)
    rt_set_by_desa = defaultdict(set)
    rw_rt_by_desa = defaultdict(lambda: {'RW': set(), 'RT': set()})
    matched_true = 0
    matched_false = 0

    def normalize_row(row: list, idx: dict[str, int], source_tab: str) -> dict:
        def get(name: str):
            position = idx.get(name)
            return safe_value(row[position]) if position is not None and position < len(row) else None

        def first_available(*names: str):
            for name in names:
                value = get(name)
                if value not in (None, ''):
                    return value
            return None

        kecamatan = get('nmkec')
        desa = get('nmdesa')
        nmsls = get('nmsls')
        rt, rw = parse_rw_rt(nmsls)
        match_status = 'MATCH' if str(get('is_matched')).upper() == 'TRUE' else 'NOT_MATCH'
        lat_raw = as_float(first_available(
            'source_latitude_normalized',
            'source_latitude',
            'scraping_lat_x',
            'match_latitude',
            'candidate_latitude',
        ))
        lon_raw = as_float(first_available(
            'source_longitude_normalized',
            'source_longitude',
            'scraping_lon_x',
            'match_longitude',
            'candidate_longitude',
        ))
        lat_final, lon_final = normalize_coords(lat_raw, lon_raw)
        normalized = {
            'source_tab': source_tab,
            'id_scraping': get('id_scraping'),
            'nama_usaha_sumber': get('nama_usaha_sumber'),
            'kategori_sumber': get('kategori_sumber'),
            'subkategori_sumber': get('subkategori_sumber'),
            'nmkec': kecamatan,
            'nmdesa': desa,
            'nmsls': nmsls,
            'rt': rt,
            'rw': rw,
            'is_matched': str(get('is_matched')).upper() == 'TRUE',
            'match_status': match_status,
            'source_latitude_normalized': lat_raw,
            'source_longitude_normalized': lon_raw,
            'source_latitude': lat_final,
            'source_longitude': lon_final,
            'match_idsbr': get('match_idsbr'),
            'match_nama_usaha': get('match_nama_usaha'),
            'similarity_score': get('similarity_score'),
            'jarak_km': get('jarak_km'),
            'keterangan': get('keterangan'),
        }
        return normalized

    for row in gmaps_ws.iter_rows(min_row=2, values_only=True):
        normalized = normalize_row(list(row), gmaps_idx, 'Master_GoogleMaps')
        gmaps_rows.append(normalized)
        if normalized['nmkec']:
            kecamatan_set.add(normalized['nmkec'])
            if normalized['nmdesa']:
                desa_set.add(normalized['nmdesa'])
                desa_by_kecamatan[normalized['nmkec']].add(normalized['nmdesa'])
                if normalized['rw']:
                    rw_rt_by_desa[normalized['nmdesa']]['RW'].add(normalized['rw'])
                if normalized['rt']:
                    rw_rt_by_desa[normalized['nmdesa']]['RT'].add(normalized['rt'])
        if normalized['is_matched']:
            matched_true += 1
        else:
            matched_false += 1

    for row in tokped_ws.iter_rows(min_row=2, values_only=True):
        normalized = normalize_row(list(row), tokped_idx, 'Master_Tokopedia')
        tokped_rows.append(normalized)
        if normalized['nmkec']:
            kecamatan_set.add(normalized['nmkec'])
            if normalized['nmdesa']:
                desa_set.add(normalized['nmdesa'])
                desa_by_kecamatan[normalized['nmkec']].add(normalized['nmdesa'])
                if normalized['rw']:
                    rw_rt_by_desa[normalized['nmdesa']]['RW'].add(normalized['rw'])
                if normalized['rt']:
                    rw_rt_by_desa[normalized['nmdesa']]['RT'].add(normalized['rt'])
        if normalized['is_matched']:
            matched_true += 1
        else:
            matched_false += 1

    preview_cards = gmaps_rows[:5] + tokped_rows[:5]
    cards = gmaps_rows + tokped_rows

    usaha_besar_rows = []
    if 'Data_Usaha_Besar' in wb.sheetnames:
        ub_ws = wb['Data_Usaha_Besar']
        ub_header = list(next(ub_ws.iter_rows(min_row=1, max_row=1, values_only=True), []))
        ub_idx = {name: i for i, name in enumerate(ub_header) if name}

        for row in ub_ws.iter_rows(min_row=2, values_only=True):
            row = list(row)
            if not any(value not in (None, '') for value in row):
                continue

            def ub_get(name: str):
                position = ub_idx.get(name)
                return safe_value(row[position]) if position is not None and position < len(row) else None

            usaha_besar_rows.append({
                'id_usaha_besar': ub_get('id_usaha_besar'),
                'nama_usaha': ub_get('nama_usaha'),
                'nama_pencacah': ub_get('nama_pencacah'),
                'status': ub_get('status'),
            })

    kbli_rows = []
    if 'Daftar_KBLI' in wb.sheetnames:
        kbli_ws = wb['Daftar_KBLI']
        kbli_header = list(next(kbli_ws.iter_rows(min_row=1, max_row=1, values_only=True), []))
        kbli_idx = {name: i for i, name in enumerate(kbli_header) if name}

        for row in kbli_ws.iter_rows(min_row=2, values_only=True):
            row = list(row)
            if not any(value not in (None, '') for value in row):
                continue

            def kbli_get(name: str):
                position = kbli_idx.get(name)
                return safe_value(row[position]) if position is not None and position < len(row) else None

            kbli_rows.append({
                'kode': kbli_get('Kode') or kbli_get('KBLI'),
                'deskripsi': kbli_get('Deskripsi') or kbli_get('Judul'),
            })

    seed = {
        'meta': {
            'source_file': WORKBOOK_PATH.name,
            'generated_at': datetime.now().isoformat(timespec='seconds'),
            'ui_ready': True,
            'sheets': ['Master_GoogleMaps', 'Master_Tokopedia', 'Data_Usaha_Besar', 'Daftar_KBLI', 'Hasil_Verifikasi'],
            'total_rows': len(gmaps_rows) + len(tokped_rows),
            'googlemaps_rows': len(gmaps_rows),
            'tokopedia_rows': len(tokped_rows),
            'kecamatan_count': len(kecamatan_set),
            'desa_count_total': len(desa_set),
            'matched_true': matched_true,
            'matched_false': matched_false,
            'usaha_besar_rows': len(usaha_besar_rows),
            'kbli_rows': len(kbli_rows),
        },
        'filters': {
            'kecamatan': sorted(kecamatan_set),
            'desa_by_kecamatan': {key: sorted(values) for key, values in sorted(desa_by_kecamatan.items())},
            'rw_rt_by_desa': {key: {'RW': sorted(values['RW']), 'RT': sorted(values['RT'])} for key, values in sorted(rw_rt_by_desa.items())},
        },
        'ui_state_shape': {
            'selected_source_tab': 'Master_GoogleMaps | Master_Tokopedia | ALL',
            'selected_kecamatan': 'string | null',
            'selected_desa': 'string | null',
            'selected_rw': 'string | null',
            'selected_rt': 'string | null',
            'search_query': 'string | null',
            'match_status': 'MATCH | NOT_MATCH | ALL',
            'cards': 'array of business card objects',
        },
        'cards': cards,
        'preview_cards': preview_cards,
        'usaha_besar': usaha_besar_rows,
        'kbli': kbli_rows,
    }
    return seed


def main() -> None:
    seed = build_seed()
    OUTPUT_PATH.write_text(json.dumps(seed, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'Wrote {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
