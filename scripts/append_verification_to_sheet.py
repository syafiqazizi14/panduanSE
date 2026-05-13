from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

COLUMNS = [
    "timestamp",
    "id_scraping",
    "source_tab",
    "match_idsbr",
    "match_nama_usaha",
    "match_alamat",
    "verification_status",
    "officer_name",
    "officer_id",
    "officer_latitude",
    "officer_longitude",
    "verified_latitude",
    "verified_longitude",
    "distance_km",
    "notes",
    "photo_url",
    "device_id",
    "idempotency_key",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Append verification rows to Hasil_Verifikasi sheet.")
    parser.add_argument("--input", required=True, help="Path to JSON file containing payload rows.")
    parser.add_argument("--workbook", required=True, help="Path to workbook XLSX.")
    parser.add_argument("--sheet", default="Hasil_Verifikasi", help="Target sheet name.")
    return parser.parse_args()


def normalize(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        s = v.strip()
        return s if s != "" else None
    return str(v)


def ensure_workbook(path: Path):
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    return wb


def ensure_sheet_and_header(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    if ws.max_row < 1:
        ws.append(COLUMNS)
        return ws

    first_row = [ws.cell(row=1, column=i + 1).value for i in range(len(COLUMNS))]
    if all(value is None for value in first_row):
        ws.delete_rows(1, 1)
        ws.append(COLUMNS)
    elif [normalize(v) for v in first_row] != COLUMNS:
        # Keep existing workbook content intact; only enforce required headers if empty.
        pass

    return ws


def get_existing_idempotency_keys(ws) -> set[str]:
    keys = set()

    header_index = None
    for idx, value in enumerate([ws.cell(row=1, column=i + 1).value for i in range(ws.max_column)], start=1):
        if normalize(value) == "idempotency_key":
            header_index = idx
            break

    if header_index is None:
        return keys

    for row in range(2, ws.max_row + 1):
        key = normalize(ws.cell(row=row, column=header_index).value)
        if key:
            keys.add(str(key))

    return keys


def to_iso_timestamp(value: str | None) -> str:
    if value:
        return value
    return datetime.now().isoformat(timespec="seconds")


def main() -> None:
    args = parse_args()

    input_path = Path(args.input)
    workbook_path = Path(args.workbook)

    rows = json.loads(input_path.read_text(encoding="utf-8"))
    if not isinstance(rows, list):
        raise ValueError("Input JSON must be an array")

    wb = ensure_workbook(workbook_path)
    ws = ensure_sheet_and_header(wb, args.sheet)

    existing_keys = get_existing_idempotency_keys(ws)
    appended_keys: list[str] = []
    skipped_existing_keys: list[str] = []

    for row in rows:
        if not isinstance(row, dict):
            continue

        idem = normalize(row.get("idempotency_key"))
        if not idem:
            continue

        if idem in existing_keys:
            skipped_existing_keys.append(str(idem))
            continue

        output_row = [
            to_iso_timestamp(normalize(row.get("timestamp"))),
            normalize(row.get("id_scraping")),
            normalize(row.get("source_tab")),
            normalize(row.get("match_idsbr")),
            normalize(row.get("match_nama_usaha")),
            normalize(row.get("match_alamat")),
            normalize(row.get("verification_status")),
            normalize(row.get("officer_name")),
            normalize(row.get("officer_id")),
            normalize(row.get("officer_latitude")),
            normalize(row.get("officer_longitude")),
            normalize(row.get("verified_latitude")),
            normalize(row.get("verified_longitude")),
            normalize(row.get("distance_km")),
            normalize(row.get("notes")),
            normalize(row.get("photo_url")),
            normalize(row.get("device_id")),
            str(idem),
        ]

        ws.append(output_row)
        existing_keys.add(str(idem))
        appended_keys.append(str(idem))

    wb.save(workbook_path)

    print(
        json.dumps(
            {
                "ok": True,
                "workbook": str(workbook_path),
                "sheet": args.sheet,
                "appended": len(appended_keys),
                "skipped_existing": len(skipped_existing_keys),
                "appended_keys": appended_keys,
                "skipped_existing_keys": skipped_existing_keys,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
