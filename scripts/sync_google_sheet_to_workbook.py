from __future__ import annotations

import argparse
import json
from pathlib import Path

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from openpyxl import Workbook


DEFAULT_SHEETS = [
    'Master_GoogleMaps',
    'Master_Tokopedia',
    'Data_Usaha_Besar',
    'Daftar_KBLI',
    'Hasil_Verifikasi',
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description='Sync Google Spreadsheet tabs into local XLSX workbook using service account.'
    )
    parser.add_argument('--spreadsheet-id', required=True, help='Google Spreadsheet ID')
    parser.add_argument('--credentials', required=True, help='Path to service account JSON file')
    parser.add_argument('--output', required=True, help='Path output XLSX workbook')
    parser.add_argument(
        '--sheets',
        nargs='*',
        default=DEFAULT_SHEETS,
        help='Sheet names to sync (space separated)',
    )
    return parser.parse_args()


def normalize_rows(values: list[list[str]]) -> list[list[str]]:
    if not values:
        return []
    max_len = max(len(row) for row in values)
    normalized = []
    for row in values:
        if len(row) < max_len:
            normalized.append(row + [''] * (max_len - len(row)))
        else:
            normalized.append(row)
    return normalized


def main() -> None:
    args = parse_args()

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    creds = Credentials.from_service_account_file(
        args.credentials,
        scopes=['https://www.googleapis.com/auth/spreadsheets.readonly'],
    )
    service = build('sheets', 'v4', credentials=creds, cache_discovery=False)
    sheets_api = service.spreadsheets().values()

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    summary: dict[str, int] = {}
    skipped: list[str] = []

    for sheet_name in args.sheets:
        range_name = f"'{sheet_name}'"
        try:
            res = sheets_api.get(
                spreadsheetId=args.spreadsheet_id,
                range=range_name,
                majorDimension='ROWS',
            ).execute()
            values = res.get('values', [])
        except Exception:
            skipped.append(sheet_name)
            continue

        ws = wb.create_sheet(title=sheet_name)
        rows = normalize_rows(values)
        for row in rows:
            ws.append(row)
        summary[sheet_name] = len(rows)

    if not wb.sheetnames:
        raise RuntimeError('No sheets were synced. Check spreadsheet ID, sharing permission, and sheet names.')

    wb.save(output_path)

    print(
        json.dumps(
            {
                'ok': True,
                'spreadsheet_id': args.spreadsheet_id,
                'output': str(output_path),
                'synced_sheets': summary,
                'skipped_sheets': skipped,
            },
            ensure_ascii=False,
        )
    )


if __name__ == '__main__':
    main()
